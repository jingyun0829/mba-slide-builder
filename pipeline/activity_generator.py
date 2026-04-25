"""Build an .xlsx file from an activity's excel_spec.

The excel_spec comes from the outline generator. Each sheet is either:
  - tabular "data": [[row], [row], ...] with the first row as headers, OR
  - keyed "cells": list of {"cell": "A1", "value": "...", "formula": "...", "bold": bool, "size": int}

We use openpyxl's basic features so the output opens cleanly in any modern
Excel version (including older ones that don't support dynamic arrays).
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_INPUT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
_SOLUTION_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")
_THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _apply_cell_style(cell, spec: dict):
    """Apply bold/size/fill/alignment hints from a cell spec dict."""
    bold = bool(spec.get("bold"))
    size = int(spec.get("size", 11))
    cell.font = Font(bold=bold, size=size, name="Calibri")

    fill_hint = (spec.get("fill") or "").lower()
    if fill_hint == "header":
        cell.fill = _HEADER_FILL
    elif fill_hint == "input":
        cell.fill = _INPUT_FILL
    elif fill_hint == "solution":
        cell.fill = _SOLUTION_FILL
    elif fill_hint and fill_hint.startswith("#"):
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_hint.lstrip("#"))


def _render_tabular_sheet(ws, data: list[list]):
    if not data:
        return
    for row_idx, row in enumerate(data, 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                # Header styling
                cell.font = Font(bold=True, size=11, name="Calibri")
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
                cell.border = _THIN_BORDER
            else:
                cell.border = _THIN_BORDER

    # Auto-width columns based on header length (rough heuristic)
    for col_idx, header in enumerate(data[0], 1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(header)) if header is not None else 8
        # find max length in this column
        max_len = header_len
        for row in data[1:]:
            if col_idx - 1 < len(row):
                v = row[col_idx - 1]
                if v is not None:
                    max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 28)


def _render_cells_sheet(ws, cells: list[dict]):
    for spec in cells or []:
        if not isinstance(spec, dict):
            continue
        addr = spec.get("cell")
        if not addr:
            continue
        try:
            cell = ws[addr]
        except Exception:
            continue

        if "formula" in spec and spec["formula"]:
            f = spec["formula"]
            if not f.startswith("="):
                f = "=" + f
            cell.value = f
        elif "value" in spec:
            cell.value = spec["value"]

        _apply_cell_style(cell, spec)

        if spec.get("note"):
            try:
                from openpyxl.comments import Comment
                cell.comment = Comment(spec["note"], "Instructor")
            except Exception:
                pass


def generate_activity_xlsx(excel_spec: dict, output_path: str) -> str:
    wb = Workbook()
    # Remove the auto-created blank sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    sheets = (excel_spec or {}).get("sheets") or []
    if not sheets:
        raise ValueError("excel_spec has no sheets to render")

    for sheet_spec in sheets:
        name = (sheet_spec.get("name") or "Sheet")[:31]
        ws = wb.create_sheet(title=name)

        if "data" in sheet_spec and sheet_spec["data"]:
            _render_tabular_sheet(ws, sheet_spec["data"])
        elif "cells" in sheet_spec:
            _render_cells_sheet(ws, sheet_spec["cells"])

    # Prepend a scenario / instructions sheet if present
    scenario = (excel_spec or {}).get("scenario")
    if scenario:
        intro = wb.create_sheet(title="Scenario", index=0)
        intro["A1"] = "Warm-up activity"
        intro["A1"].font = Font(bold=True, size=16)
        intro["A3"] = scenario
        intro["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        intro.column_dimensions["A"].width = 100
        intro.row_dimensions[3].height = 80

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
